import pandas as pd
import pyodbc
import datetime
from openpyxl import Workbook


cnxn = pyodbc.connect("DSN=L5QNS_PROD;UID=;DBQ=TNS:L5QNS_PROD.WORLD;ASY=OFF")
cursor = cnxn.cursor()

section_list = []
section = {}
section['Unit'] = 'MWEGL'
section['id_audits'] = []
section['coil_slab'] = []
section_list.append(section)
section = {}
section['Unit'] = 'MWAL'
section['id_audits'] = []
section['coil_slab'] = []
section_list.append(section)
section = {}
section['Unit'] = 'MWSA'
section['id_audits'] = []
section['coil_slab'] = []
section_list.append(section)

today = datetime.date.today()
idx = (today.weekday() + 1) % 7
sun = today - datetime.timedelta(7 + idx)
sat = today - datetime.timedelta(7 + idx - 7)

sat_date = today - datetime.timedelta(7 + idx - 6)


for section in section_list:
    verification_audit_sql = """SELECT QUALITY_OBSERVATIONS.SC_NUMBER, QUALITY_OBSERVATIONS.EMPLOYEE_ID, QUALITY_OBSERVATIONS.EMPLOYEE_FIRST_NAME, QUALITY_OBSERVATIONS.EMPLOYEE_LAST_NAME, QUALITY_OBSERVATIONS.HOME_DEPT_ID, QUALITY_OBSERVATIONS.EMPLOYEE_STATUS_CODE, QUALITY_OBSERVATIONS.HOME_DEPT_NAME, QUALITY_OBSERVATIONS.STATUS_DESC, QUALITY_OBSERVATIONS.OBSERVED_DATE, QUALITY_OBSERVATIONS.DAILY_SEQ_NUMBER, QUALITY_OBSERVATIONS.OBSERVED_DEPT_ID, QUALITY_OBSERVATIONS.OBS_TOPIC, QUALITY_OBSERVATIONS.OBS_CATEGORY_ID, QUALITY_OBSERVATIONS.CORRECTIVE_ACTION_ID, QUALITY_OBSERVATIONS.OBS_CATEGORY_TYPE_ID, QUALITY_OBSERVATIONS.OBS_BY_ID, QUALITY_OBSERVATIONS.INITIATED_BY_INCIDENT_IND, QUALITY_OBSERVATIONS.OBSERVED_REVIEW_DOC_TYPE_DESC, QUALITY_OBSERVATIONS.OBSERVED_DOC_TITLE, QUALITY_OBSERVATIONS.DOC_TYPE_NUMBER, QUALITY_OBSERVATIONS.OBS_NATURE_ID, QUALITY_OBSERVATIONS.CATEGORY_DESC, QUALITY_OBSERVATIONS.CORRECTIVE_ACTION_DESC, QUALITY_OBSERVATIONS.OBS_CATEGORY_TYPE_DESC, QUALITY_OBSERVATIONS.OBSERVER_FIRST_NAME, QUALITY_OBSERVATIONS.OBSERVER_LAST_NAME, QUALITY_OBSERVATIONS.OBS_NATURE_DESC, QUALITY_OBSERVATIONS.DOC_TITLE, QUALITY_OBSERVATIONS.REVIEW_DOC_TYPE_ID, QUALITY_OBSERVATIONS.REVIEW_DOC_TYPE_DESC, QUALITY_OBSERVATIONS.GROUP_ID, QUALITY_OBSERVATIONS.GROUP_NAME, QUALITY_OBSERVATIONS.PLANT_ID, QUALITY_OBSERVATIONS.PLANT_NAME, QUALITY_OBSERVATIONS.OBSERVED_DEPT_NAME, QUALITY_OBSERVATIONS.OBSERVED_GROUP_ID, QUALITY_OBSERVATIONS.OBSERVED_GROUP_NAME, QUALITY_OBSERVATIONS.DATA_ENTRY_TYPE_CODE, QUALITY_OBSERVATIONS.TARGETED_OBSERVATION_IND, QUALITY_OBSERVATIONS.OBS_GRADE_ID, QUALITY_OBSERVATIONS.OBS_GRADE_VALUE, QUALITY_OBSERVATIONS.OBS_GRADE_DESC
    FROM SCDB.QUALITY_OBSERVATIONS QUALITY_OBSERVATIONS
    WHERE (QUALITY_OBSERVATIONS.HOME_DEPT_ID='""" + str(section['Unit']) + """') AND (QUALITY_OBSERVATIONS.OBS_CATEGORY_TYPE_DESC='Identity Verification Audit') AND (QUALITY_OBSERVATIONS.OBSERVED_DATE Between {ts '""" + str(sun) + """ 00:00:00'} And {ts '""" + str(sat) + """ 00:00:00'})
    ORDER BY QUALITY_OBSERVATIONS.OBSERVED_DATE DESC"""


    verification_audits = pd.read_sql(verification_audit_sql, cnxn)
    id_audits_list = verification_audits.T.to_dict().values()

    coil_slab_sql = """SELECT QUALITY_OBSERVATIONS.SC_NUMBER, QUALITY_OBSERVATIONS.EMPLOYEE_ID, QUALITY_OBSERVATIONS.EMPLOYEE_FIRST_NAME, QUALITY_OBSERVATIONS.EMPLOYEE_LAST_NAME, QUALITY_OBSERVATIONS.HOME_DEPT_ID, QUALITY_OBSERVATIONS.EMPLOYEE_STATUS_CODE, QUALITY_OBSERVATIONS.HOME_DEPT_NAME, QUALITY_OBSERVATIONS.STATUS_DESC, QUALITY_OBSERVATIONS.OBSERVED_DATE, QUALITY_OBSERVATIONS.DAILY_SEQ_NUMBER, QUALITY_OBSERVATIONS.OBSERVED_DEPT_ID, QUALITY_OBSERVATIONS.OBS_TOPIC, QUALITY_OBSERVATIONS.OBS_CATEGORY_ID, QUALITY_OBSERVATIONS.CORRECTIVE_ACTION_ID, QUALITY_OBSERVATIONS.OBS_CATEGORY_TYPE_ID, QUALITY_OBSERVATIONS.OBS_BY_ID, QUALITY_OBSERVATIONS.INITIATED_BY_INCIDENT_IND, QUALITY_OBSERVATIONS.OBSERVED_REVIEW_DOC_TYPE_DESC, QUALITY_OBSERVATIONS.OBSERVED_DOC_TITLE, QUALITY_OBSERVATIONS.DOC_TYPE_NUMBER, QUALITY_OBSERVATIONS.OBS_NATURE_ID, QUALITY_OBSERVATIONS.CATEGORY_DESC, QUALITY_OBSERVATIONS.CORRECTIVE_ACTION_DESC, QUALITY_OBSERVATIONS.OBS_CATEGORY_TYPE_DESC, QUALITY_OBSERVATIONS.OBSERVER_FIRST_NAME, QUALITY_OBSERVATIONS.OBSERVER_LAST_NAME, QUALITY_OBSERVATIONS.OBS_NATURE_DESC, QUALITY_OBSERVATIONS.DOC_TITLE, QUALITY_OBSERVATIONS.REVIEW_DOC_TYPE_ID, QUALITY_OBSERVATIONS.REVIEW_DOC_TYPE_DESC, QUALITY_OBSERVATIONS.GROUP_ID, QUALITY_OBSERVATIONS.GROUP_NAME, QUALITY_OBSERVATIONS.PLANT_ID, QUALITY_OBSERVATIONS.PLANT_NAME, QUALITY_OBSERVATIONS.OBSERVED_DEPT_NAME, QUALITY_OBSERVATIONS.OBSERVED_GROUP_ID, QUALITY_OBSERVATIONS.OBSERVED_GROUP_NAME, QUALITY_OBSERVATIONS.DATA_ENTRY_TYPE_CODE, QUALITY_OBSERVATIONS.TARGETED_OBSERVATION_IND, QUALITY_OBSERVATIONS.OBS_GRADE_ID, QUALITY_OBSERVATIONS.OBS_GRADE_VALUE, QUALITY_OBSERVATIONS.OBS_GRADE_DESC
    FROM SCDB.QUALITY_OBSERVATIONS QUALITY_OBSERVATIONS
    WHERE (QUALITY_OBSERVATIONS.HOME_DEPT_ID='""" + str(section['Unit']) + """') AND (QUALITY_OBSERVATIONS.OBS_CATEGORY_TYPE_DESC='Coil/Slab Identity') AND (QUALITY_OBSERVATIONS.OBSERVED_DATE Between {ts '""" + str(sun) + """ 00:00:00'} And {ts '""" + str(sat) + """ 00:00:00'})
    ORDER BY QUALITY_OBSERVATIONS.OBSERVED_DATE DESC"""

    csi = pd.read_sql(coil_slab_sql, cnxn)
    csi_list = csi.T.to_dict().values()

    section['id_audits'] = id_audits_list
    section['coil_slab'] = csi_list

wb = Workbook()
file_name = sat_date.strftime('%m%d%y') + "_id_observations.xlsx"
dest_filename = file_name
for idx, section in enumerate(section_list):
    if idx == 0:
        ws = wb.active
        ws.title = str(section['Unit'])[2:] + '_ID_Audit'
    else:
        ws = wb.create_sheet(str(section['Unit'])[2:] + '_ID_Audit')
    if len(section['id_audits']) > 0:
        header_row = 1
        ws.cell(column=1, row=header_row, value="Department")
        ws.cell(column=2, row=header_row, value="Employee Name")
        ws.cell(column=3, row=header_row, value="Obs Date")
        ws.cell(column=4, row=header_row, value="Obs Category")
        ws.cell(column=5, row=header_row, value="Obs Cat Type")
        ws.cell(column=6, row=header_row, value="Corr Actions")
        ws.cell(column=7, row=header_row, value="Observed By")
        ws.cell(column=8, row=header_row, value="Doc Type")
        ws.cell(column=9, row=header_row, value="Doc Title")
        ws.cell(column=10, row=header_row, value="Comments")
        for idx, val in enumerate(section['id_audits']):
            ws.cell(column=1, row=idx + 2, value=str(val['HOME_DEPT_NAME'].strip()))
            ws.cell(column=2, row=idx + 2, value=(str(val['EMPLOYEE_FIRST_NAME'].strip() + " " + str(val['EMPLOYEE_LAST_NAME'].strip()))))
            ws.cell(column=3, row=idx + 2, value=str(val['OBSERVED_DATE']))
            ws.cell(column=4, row=idx + 2, value=str(val['CATEGORY_DESC'].strip()))
            ws.cell(column=5, row=idx + 2, value=str(val['OBS_CATEGORY_TYPE_DESC'].strip()))
            ws.cell(column=6, row=idx + 2, value=str(val['CORRECTIVE_ACTION_DESC'].strip()))
            ws.cell(column=7, row=idx + 2, value=(str(val['OBSERVER_FIRST_NAME'].strip() + " " + str(val['OBSERVER_LAST_NAME'].strip()))))
            ws.cell(column=8, row=idx + 2, value=str(val['OBSERVED_REVIEW_DOC_TYPE_DESC'].strip()))
            ws.cell(column=9, row=idx + 2, value=str(val['OBSERVED_DOC_TITLE'].strip()))
            ws.cell(column=10, row=idx + 2, value=str(val['OBS_TOPIC'].strip()))
    else:
        ws.cell(column=1, row=1, value="Nothing Found")

for idx, section in enumerate(section_list):
    ws = wb.create_sheet(str(section['Unit'])[2:] + '_Coils_Slab_ID')
    if len(section['coil_slab']) > 0:
        header_row = 1
        ws.cell(column=1, row=header_row, value="Department")
        ws.cell(column=2, row=header_row, value="Employee Name")
        ws.cell(column=3, row=header_row, value="Obs Date")
        ws.cell(column=4, row=header_row, value="Obs Category")
        ws.cell(column=5, row=header_row, value="Obs Cat Type")
        ws.cell(column=6, row=header_row, value="Corr Actions")
        ws.cell(column=7, row=header_row, value="Observed By")
        ws.cell(column=8, row=header_row, value="Doc Type")
        ws.cell(column=9, row=header_row, value="Doc Title")
        ws.cell(column=10, row=header_row, value="Comments")
        for idx, val in enumerate(section['coil_slab']):
            ws.cell(column=1, row=idx + 2, value=str(val['HOME_DEPT_NAME'].strip()))
            ws.cell(column=2, row=idx + 2, value=(str(val['EMPLOYEE_FIRST_NAME'].strip() + " " + str(val['EMPLOYEE_LAST_NAME'].strip()))))
            ws.cell(column=3, row=idx + 2, value=str(val['OBSERVED_DATE']))
            ws.cell(column=4, row=idx + 2, value=str(val['CATEGORY_DESC'].strip()))
            ws.cell(column=5, row=idx + 2, value=str(val['OBS_CATEGORY_TYPE_DESC'].strip()))
            ws.cell(column=6, row=idx + 2, value=str(val['CORRECTIVE_ACTION_DESC'].strip()))
            ws.cell(column=7, row=idx + 2, value=(str(val['OBSERVER_FIRST_NAME'].strip() + " " + str(val['OBSERVER_LAST_NAME'].strip()))))
            ws.cell(column=8, row=idx + 2, value=str(val['OBSERVED_REVIEW_DOC_TYPE_DESC'].strip()))
            ws.cell(column=9, row=idx + 2, value=str(val['OBSERVED_DOC_TITLE'].strip()))
            ws.cell(column=10, row=idx + 2, value=str(val['OBS_TOPIC'].strip()))
    else:
        ws.cell(column=1, row=1, value="Nothing Found")

wb.save(filename = dest_filename)

import win32com.client as win32
win32c = win32.constants

excel = win32.gencache.EnsureDispatch('Excel.Application')
excel.Visible = False
try:
    wb = excel.Workbooks.Open('C:\\Users\\a10965\\PycharmProjects\\morning_report3\\' + file_name)
except:
    print "Failed to open spreadsheet ABCDCatering.xls"
    sys.exit(1)

worksheets = excel.ActiveWorkbook.Sheets
for s in worksheets:
    s.Select()
    s.Columns.AutoFit()
    s.Rows.AutoFit()
    xldata = s.UsedRange.Value
    s.UsedRange.Select()
    s.ListObjects.Add(XlListObjectHasHeaders=win32c.xlYes).TableStyle = "TableStyleMedium15"

wb.Save()
excel.Application.Quit()

#### SEND THE EMAIL
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.MIMEMultipart import MIMEMultipart
from email.MIMEBase import MIMEBase
from email import Encoders

# me == my email address
# you == recipient's email address
me = "chris.schottmiller@aksteel.com"
you = ["chris.schottmiller@aksteel.com", "Sal.Ciriello@aksteel.com", "TonyL.Webb@aksteel.com", "Glenn.Stamper@aksteel.com", "Chris.Fulks@aksteel.com", "Greg.Robertson@aksteel.com", "Scott.Beckerman@aksteel.com", "Antonio.Carter@aksteel.com", "Gary.Larson@aksteel.com"]
#you = "chris.schottmiller@aksteel.com, sal.ciriello@aksteel.com, paul.reed@akstee.com, chris.fulks@aksteel.com, glenn.stamper@aksteel.com"
# Create message container - the correct MIME type is multipart/alternative.
msg = MIMEMultipart('alternative')
msg['Subject'] = "Identity Observations Week Ending " + sat_date.strftime('%m%d%y')
msg['From'] = me
msg['To'] = ", ".join(you)
# Create the body of the message (a plain-text and an HTML version).
text = "Identity Observations Update"
html = """\
<html>
  <head></head>
  <body>
    <p>South Processing Identity Quality Observations for the week ending """ + sat_date.strftime('%m%d%y') + """<br>
       <br>
       Click <a href="mailto:Chris.Schottmiller@aksteel.com?Subject=Weekly%20Quality%20Email%20Issue">here</a> and email to request changes..
    </p>
  </body>
</html>
"""
# Record the MIME types of both parts - text/plain and text/html.
part1 = MIMEText(text, 'plain')
part2 = MIMEText(html, 'html')
# Attach parts into message container.
# According to RFC 2046, the last part of a multipart message, in this case
# the HTML message, is best and preferred.
msg.attach(part1)
msg.attach(part2)

part = MIMEBase('application', 'octet-stream')
part.set_payload(open(file_name, 'rb').read())
Encoders.encode_base64(part)

part.add_header('Content-Disposition', 'attachment; filename="' + file_name +'"')
msg.attach(part)
# Send the message via local SMTP server.
s = smtplib.SMTP('smtp.akst.com')
# sendmail function takes 3 arguments: sender's address, recipient's address
# and message to send - here it is sent as one string.
s.sendmail(me, you, msg.as_string())
s.quit()
